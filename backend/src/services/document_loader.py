"""
Universal Document Loader.
Automatically detects and extracts text from various file types.
"""

import os
from pathlib import Path
from typing import Optional, Dict, Any

SUPPORTED_EXTENSIONS = {
    ".txt": "text",
    ".md": "markdown",
    ".pdf": "pdf",
    ".docx": "docx",
    ".doc": "docx",
    ".json": "json",
    ".csv": "csv",
    ".xlsx": "excel",
    ".xls": "excel",
    ".yaml": "yaml",
    ".yml": "yaml",
    ".xml": "xml",
    ".html": "html",
    ".htm": "html",
    ".py": "code",
    ".js": "code",
    ".ts": "code",
    ".jsx": "code",
    ".tsx": "code",
    ".java": "code",
    ".cpp": "code",
    ".c": "code",
    ".h": "code",
    ".rs": "code",
    ".go": "code",
    ".rb": "code",
    ".php": "code",
    ".swift": "code",
    ".kt": "code",
    ".sql": "code",
    ".sh": "code",
    ".bash": "code",
    ".css": "code",
    ".scss": "code",
    ".less": "code",
    ".vue": "code",
    ".svelte": "code",
    ".r": "code",
    ".scala": "code",
    ".lua": "code",
    ".perl": "code",
    ".pl": "code",
    ".png": "image",
    ".jpg": "image",
    ".jpeg": "image",
    ".gif": "image",
    ".bmp": "image",
    ".tiff": "image",
    ".webp": "image",
    ".mp3": "audio",
    ".wav": "audio",
    ".m4a": "audio",
    ".flac": "audio",
    ".ogg": "audio",
    ".aac": "audio",
    ".wma": "audio",
    ".mp4": "video",
    ".mkv": "video",
    ".avi": "video",
    ".mov": "video",
    ".webm": "video",
}

# Human-readable labels for UI
FILE_TYPE_LABELS = {
    "text": "Text Document",
    "markdown": "Markdown",
    "pdf": "PDF Document",
    "docx": "Word Document",
    "json": "JSON File",
    "csv": "CSV Spreadsheet",
    "excel": "Excel Spreadsheet",
    "yaml": "YAML File",
    "xml": "XML File",
    "html": "HTML Page",
    "code": "Source Code",
    "image": "Image",
    "audio": "Audio",
    "video": "Video",
}


def get_file_type(file_path: str) -> Optional[str]:
    """Get the file type category from extension."""
    ext = Path(file_path).suffix.lower()
    return SUPPORTED_EXTENSIONS.get(ext)


def get_supported_extensions() -> list:
    """Return list of all supported file extensions."""
    return list(SUPPORTED_EXTENSIONS.keys())


def load_text(file_path: str) -> str:
    """Load plain text file."""
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def load_markdown(file_path: str) -> str:
    """Load markdown file (treated as text)."""
    return load_text(file_path)


def load_pdf(file_path: str) -> str:
    """Load PDF file using pypdf, with OCR fallback for scanned docs."""
    from pypdf import PdfReader

    text_parts = []
    try:
        reader = PdfReader(file_path)
        for page in reader.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)
    except Exception as e:
        print(f"Error reading PDF text natively: {e}")

    extracted_text = "\n\n".join(text_parts)

    if len(extracted_text.strip()) < 100:
        print(f"PDF {os.path.basename(file_path)} appears scanned (text len < 100). Attempting OCR...")
        try:
            from pdf2image import convert_from_path
            import pytesseract

            try:
                pytesseract.get_tesseract_version()
            except Exception:
                print("Tesseract not installed. Skipping OCR.")
                return extracted_text

            images = convert_from_path(file_path)
            ocr_parts = []
            for i, img in enumerate(images):
                print(f"OCR processing page {i+1}/{len(images)}...")
                text = pytesseract.image_to_string(img)
                ocr_parts.append(text)

            return "\n\n".join(ocr_parts)

        except ImportError:
            print("pdf2image or pytesseract missing. Install them for OCR support.")
        except Exception as e:
            print(f"OCR failed for PDF: {e}")

    return extracted_text


def load_docx(file_path: str) -> str:
    """Load Word document using python-docx."""
    try:
        import docx
        doc = docx.Document(file_path)
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
        return "\n\n".join(paragraphs)
    except ImportError:
        raise ImportError("python-docx is required for .docx files: pip install python-docx")


def load_json(file_path: str) -> str:
    """Load JSON file and convert to readable text."""
    import json
    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return json.dumps(data, indent=2, ensure_ascii=False)


def load_csv(file_path: str) -> str:
    """Load CSV file and convert to readable text."""
    import csv
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return ""
    lines = []
    lines.append("| " + " | ".join(rows[0]) + " |")
    lines.append("| " + " | ".join(["---"] * len(rows[0])) + " |")
    for row in rows[1:]:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def load_excel(file_path: str) -> str:
    """Load Excel file and convert to readable text."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True)
        all_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_text.append(f"## Sheet: {sheet_name}\n")
            for row in ws.iter_rows(values_only=True):
                all_text.append(" | ".join(str(c) if c is not None else "" for c in row))
        wb.close()
        return "\n".join(all_text)
    except ImportError:
        raise ImportError("openpyxl is required for .xlsx files: pip install openpyxl")


def load_yaml(file_path: str) -> str:
    """Load YAML file and convert to readable text."""
    try:
        import yaml
        with open(file_path, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)
        return yaml.dump(data, default_flow_style=False, allow_unicode=True)
    except ImportError:
        raise ImportError("PyYAML is required for .yaml files: pip install pyyaml")


def load_html(file_path: str) -> str:
    """Load HTML file and extract text content."""
    try:
        from bs4 import BeautifulSoup
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            soup = BeautifulSoup(f, "html.parser")
        for script in soup(["script", "style"]):
            script.decompose()
        return soup.get_text(separator="\n", strip=True)
    except ImportError:
        raise ImportError("beautifulsoup4 is required for .html files: pip install beautifulsoup4")


def load_xml(file_path: str) -> str:
    """Load XML file and convert to readable text."""
    try:
        import xml.etree.ElementTree as ET

        tree = ET.parse(file_path)
        root = tree.getroot()

        def extract_text(element, depth=0):
            lines = []
            tag = element.tag.split("}")[-1] if "}" in element.tag else element.tag
            if element.text and element.text.strip():
                lines.append(f"{'  ' * depth}{tag}: {element.text.strip()}")
            elif len(element) == 0:
                lines.append(f"{'  ' * depth}{tag}")
            for child in element:
                lines.extend(extract_text(child, depth + 1))
            if element.tail and element.tail.strip():
                lines.append(f"{'  ' * depth}{element.tail.strip()}")
            return lines

        return "\n".join(extract_text(root))
    except Exception:
        return load_text(file_path)


def load_code(file_path: str) -> str:
    """Load source code file with language header."""
    ext = Path(file_path).suffix.lower()
    language_map = {
        ".py": "python", ".js": "javascript", ".ts": "typescript",
        ".jsx": "javascript", ".tsx": "typescript", ".java": "java",
        ".cpp": "cpp", ".c": "c", ".h": "c", ".rs": "rust",
        ".go": "go", ".rb": "ruby", ".php": "php", ".swift": "swift",
        ".kt": "kotlin", ".sql": "sql", ".sh": "bash", ".bash": "bash",
        ".css": "css", ".scss": "scss", ".less": "less", ".vue": "vue",
        ".svelte": "svelte", ".r": "r", ".scala": "scala", ".lua": "lua",
        ".perl": "perl", ".pl": "perl",
    }
    language = language_map.get(ext, "code")
    content = load_text(file_path)
    return f"```{language}\n{content}\n```"


def load_image_ocr(file_path: str) -> str:
    """Load image and extract text using OCR."""
    try:
        from PIL import Image

        try:
            import pytesseract
            pytesseract.get_tesseract_version()
        except Exception:
            image = Image.open(file_path)
            width, height = image.size
            return (
                f"[Image file: {Path(file_path).name}]\n"
                f"Dimensions: {width}x{height} pixels\n"
                f"Format: {image.format}\nMode: {image.mode}\n\n"
                f"Note: OCR not available - install tesseract-ocr to extract text from images."
            )

        image = Image.open(file_path)
        text = pytesseract.image_to_string(image)
        if text.strip():
            return text.strip()
        else:
            width, height = image.size
            return (
                f"[Image file: {Path(file_path).name}]\n"
                f"Dimensions: {width}x{height} pixels\n"
                f"No text detected in image."
            )
    except ImportError:
        raise ImportError("Pillow is required for image processing: pip install pillow")


def load_document(file_path: str) -> Dict[str, Any]:
    """
    Universal document loader.
    Automatically detects file type and extracts text.

    Returns:
        Dict with 'text', 'file_type', 'success', and optional 'error'
    """
    file_type = get_file_type(file_path)

    if file_type is None:
        return {
            "text": "",
            "file_type": "unknown",
            "success": False,
            "error": f"Unsupported file type: {Path(file_path).suffix}",
        }

    loaders = {
        "text": load_text,
        "markdown": load_markdown,
        "pdf": load_pdf,
        "docx": load_docx,
        "json": load_json,
        "csv": load_csv,
        "excel": load_excel,
        "yaml": load_yaml,
        "xml": load_xml,
        "html": load_html,
        "code": load_code,
        "image": load_image_ocr,
    }

    # Audio and video require external transcription
    if file_type in ("audio", "video"):
        return {
            "text": "",
            "file_type": file_type,
            "success": True,
            "requires_transcription": True,
        }

    loader = loaders.get(file_type)
    if loader is None:
        return {
            "text": "",
            "file_type": file_type,
            "success": False,
            "error": f"No loader available for type: {file_type}",
        }

    try:
        text = loader(file_path)
        return {
            "text": text,
            "file_type": file_type,
            "success": True,
            "char_count": len(text),
        }
    except Exception as e:
        return {
            "text": "",
            "file_type": file_type,
            "success": False,
            "error": str(e),
        }
